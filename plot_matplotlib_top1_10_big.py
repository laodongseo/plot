#‐*‐coding:utf‐8‐*‐
"""
1张大图绘制多个图片
1个图片绘制多个分类条形图

"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib
import matplotlib.pyplot as plt
import numpy as np

# 显示中文标签
plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus']=False
plt.rcParams['font.size'] = 10


# ax绘图
def draw_bar(ax,dict_like_data,city):
    # 画每个分类的条形图
    n = 0 #n辅助确定起点位置
    for key,values in dict_like_data.items():
        rects = ax.bar(Xlocation + width * n, values, width, label=key)
        ax.bar_label(rects,fmt='%g%%',padding=1)
        n+=1
    ax.set_ylabel(Ylabel)
    ax.set_title(city + bar_title)
    ax.set_xticks(Xlocation + (len(dict_like_data)-1) * 1/2 * width)
    ax.set_xticklabels(X_labels)
    ax.legend()


def main(fig_rows,fig_cols,sheet_objs):
	ImgIndex = 0 # 图片索引(sheet_objs的索引)
	for row_num in range(fig_rows):
		for col_num in range(fig_cols):
			# 循环行列,依次获取ax
			ax = Axes[row_num,col_num]
			if ImgIndex < len(sheet_objs):
				ws = sheet_objs[ImgIndex]
				sheet_dict_data = {}
				city = ws.title
				print(city)
				for domain,lis_cols in ConfigDomain.items():
					col_start,col_end = lis_cols
					for row in ws.iter_rows(min_col=col_start, max_col=col_end,min_row=last_row,max_row=last_row):
						domain_info = [i.value for i in row]
						top_percents = [ domain_info[1],domain_info[3],domain_info[5],domain_info[7] ]
						top_percents.reverse() # 改变顺序为top1-top10
						# 汇总sheet读取出来是浮点数
						top_percents = [float(str(n).replace('%','').strip())/100  for n in top_percents] if city != '汇总' else top_percents
						top_percents = [round(n*100,2) for n in top_percents]
						sheet_dict_data[domain] = top_percents
				draw_bar(ax,sheet_dict_data,city)
				ImgIndex+=1


if __name__ == "__main__":
	wb = load_workbook('bdpc小区词首页词监控top1_10分布.xlsx',data_only=True)
	sheet_objs = list(wb)
	save_path = './pc_top1_10分布图/'
	last_row = 13 #最后一行(从1开始数)
	# 配置域名所在的列数
	# ConfigDomain = {'5i5j':[3,10],'lianjia':[11,18],'ke':[19,26],'anjuke':[27,34],'fang':[35,42]}
	ConfigDomain = {'5i5j':[3,10],'lianjia':[11,18],'ke':[19,26]}
	X_labels = ['top1', 'top3', 'top5', 'top10'] # x轴标签
	Xlocation = np.arange(len(X_labels))  # the label locations
	width = 0.3  # the width of the bars
	Ylabel = '首页率'
	bar_title = '小区词排名'
	fig_rows,fig_cols = (5,4)
	Fig, Axes = plt.subplots(fig_rows,fig_cols,sharex=False,figsize=(20,20),dpi=100,facecolor='#F5F5DC')
	main(fig_rows,fig_cols,sheet_objs)
	Fig.tight_layout(h_pad=2)
	plt.savefig(f'{save_path}总1图.png')
